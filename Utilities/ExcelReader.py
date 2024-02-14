import openpyxl

path = "..//excel//Testdata.xlsx"
sheetname = "LoginTest"


def getRowcount(path, sheetname):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    return sheet.max_row


def getColcount(path, sheetname):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    return sheet.max_column


def getCellData(path, sheetname, rowNumber, colNumber):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    return sheet.cell(row=2, column=1).value


def setCellData(path, sheetname, rowNumber, colNumber, data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    sheet.cell(row=rowNumber, column=colNumber).value = data
    workbook.save(path)


rows = getRowcount(path, sheetname)
cols = getColcount(path,sheetname)

print(rows,"----",cols)

print(getCellData(path,sheetname,3,1))
setCellData(path,sheetname,1,4,"DOB")